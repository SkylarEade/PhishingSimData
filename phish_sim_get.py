import requests
import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from token_gen import get_access_token
from phishing_sim_analysis import data_clean, generate_simulation_overview

load_dotenv()
GRAPH_SCOPE = [os.getenv("GRAPH_SCOPE")]
PHISH_ENDPOINT = 'https://graph.microsoft.com/v1.0/security/attackSimulation/simulations'
USER_ENDPOINT = (
    "https://graph.microsoft.com/v1.0/users"
    "?$select=displayName,userPrincipalName,department,officeLocation"
)

def get_entra_users(headers):
    endpoint = USER_ENDPOINT
    users = []
    while endpoint:
        response = requests.get(endpoint, headers=headers)
        if response.status_code == 200:
            data = response.json()
            users.extend(data.get('value', []))
            endpoint = data.get('@odata.nextLink')
        else:
            raise Exception(f"Error {response.status_code}: {response.text}")
    return users

def get_simulations(headers):
    endpoint = PHISH_ENDPOINT
    simulations = []
    while endpoint:
        response = requests.get(endpoint, headers=headers)
        if response.status_code == 200:
            data = response.json()
            simulations.extend(data.get('value', []))
            endpoint = data.get('@odata.nextLink')
        else:
            raise Exception(f"Error {response.status_code}: {response.text}")
    return simulations

def parse_simulation_events(user_record):
    events = {e.get('eventName') for e in user_record.get('simulationEvents', [])}
    return {
        'reported': 'ReportedEmail' in events,
        'deleted': 'MessageDeleted' in events,
        'read': 'MessageRead' in events,
        'forwarded': 'MessageForwarded' in events,
        'received': 'SuccessfullyDeliveredEmail' in events
    }


def get_simulation_users(headers, simulation_id):
    endpoint = f"https://graph.microsoft.com/v1.0/security/attackSimulation/simulations/{simulation_id}/report/simulationUsers"
    users = []
    while endpoint:
        response = requests.get(endpoint, headers=headers)
        if response.status_code == 200:
            data = response.json()
            for record in data.get('value', []):
                sim_user = record.get('simulationUser', {})
                flags = parse_simulation_events(record)
                users.append({
                    'userDisplayName': sim_user.get('displayName'),
                    'userPrincipalName': sim_user.get('email'),
                    'compromised': record.get('isCompromised'),
                    'assignedTrainingsCount': record.get('assignedTrainingsCount', 0),
                    'completedTrainingsCount': record.get('completedTrainingsCount', 0),
                    **flags
                })
            endpoint = data.get('@odata.nextLink')
        else:
            raise Exception(f"Error {response.status_code}: {response.text}")
    return users

def export_xlsx(merged_rows, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    table_name = "PhishingSimTable"

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb["Phishing Simulation Data"]
        for row in merged_rows:
            ws.append(list(row.values()))
        if table_name in ws.tables:
            table = ws.tables[table_name]
            table.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Phishing Simulation Data"
        headers_row = list(merged_rows[0].keys()) if merged_rows else []
        ws.append(headers_row)
        for row in merged_rows:
            ws.append(list(row.values()))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.save(output_path)
    print(f"Data appended and table updated: {output_path}")



if __name__ == "__main__":
    try:
        graph_token = get_access_token(GRAPH_SCOPE)
        headers = {'Authorization': f'Bearer {graph_token}'}

        simulations = get_simulations(headers)
        if not simulations:
            print("No simulations found.")
            exit()

        print("\nAvailable Simulations:")
        for idx, sim in enumerate(simulations, start=1):
            sim_name = sim.get('displayName', 'Unnamed')
            sim_date = sim.get('launchDateTime', '')[:10]
            print(f"{idx}. {sim_name} ({sim_date})")

        choice = int(input("\nEnter the number of the simulation you want to use: "))
        selected_sim = simulations[choice - 1]
        sim_id = selected_sim['id']
        sim_name = selected_sim.get('displayName', 'Unnamed')
        sim_date = selected_sim.get('launchDateTime', '')[:10]

        complexity = input("Enter complexity (High, Medium, Low): ").strip().capitalize()
        if complexity not in ["High", "Medium", "Low"]:
            raise ValueError("Invalid complexity. Please enter High, Medium, or Low.")

        print(f"\nSelected Simulation: {sim_name} ({sim_date}) - Complexity: {complexity}")

        sim_users = get_simulation_users(headers, sim_id)
        print(f"Retrieved {len(sim_users)} user records for this simulation.")

        entra_users = get_entra_users(headers)
        entra_lookup = {u['userPrincipalName'].lower(): u for u in entra_users if 'userPrincipalName' in u}

        merged_rows = []
        for su in sim_users:
            principal_name = (su.get('userPrincipalName') or '').lower()
            entra_info = entra_lookup.get(principal_name, {})
            merged_rows.append({
                'Sim Name': sim_name,
                'Sim Date': sim_date,
                'Complexity': complexity,
                'User Display Name': su['userDisplayName'],
                'Compromised': su['compromised'],
                'Reported': su['reported'],
                'Deleted': su['deleted'],
                'Read': su['read'],
                'Forwarded': su['forwarded'], 
                'Received': su['received'],   
                'Trainings Assigned': su['assignedTrainingsCount'],
                'Trainings Completed': su['completedTrainingsCount'],
                'Office Location': entra_info.get('officeLocation', ''),
                'Department': entra_info.get('department', '')
            })

        export_xlsx(merged_rows, "excel/phishing_sim_data.xlsx")
        data_clean()
        generate_simulation_overview()
    except Exception as auth_error:
        print(f"Authentication failed: {auth_error}")
