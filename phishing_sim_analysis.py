from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

departments = [
    "Accounting", "Administration", "Area Management", "Assembly", "Buildings & Grounds",
    "Construction", "Customer Support", "Electrical Shop", "Engineering", "Environmental",
    "Human Resources", "IT", "Machine Shop", "Maintenance", "Management", "Production",
    "Quality", "R&D", "Safety", "Sales & Marketing", "Security", "Shipping & Receiving",
    "Site Lead", "Supply Chain"
]

locations = ["Bogor", "Henderson", "Lebanon", "Remote", "Tarui"]

def generate_simulation_overview(file_path="excel/phishing_sim_data.xlsx"):
    wb = load_workbook(file_path)
    ws_data = wb["Phishing Simulation Data"]

    headers = [cell.value for cell in ws_data[1]]
    idx = {h: i for i, h in enumerate(headers)}

    sims = {}
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        sim_name = row[idx["Sim Name"]]
        sims.setdefault(sim_name, []).append(row)

    if "Simulation Overview" in wb.sheetnames:
        wb.remove(wb["Simulation Overview"])
    ws_overview = wb.create_sheet("Simulation Overview")

    core_headers = [
        "Sim Name", "Sim Date", "Complexity",
        "Total Delivered", "Total Received", "Total Forwarded",
        "Total Read", "Total Deleted", "Total Reported",
        "Total Compromised", "% Compromised of Delivered", "% Compromised of Read"
    ]
    core_start_col = 1
    ws_overview.append(core_headers)

    dept_headers = ["Sim Name"] + [f"{d} % Comp (of Received)" for d in departments]
    dept_start_col = len(core_headers) + 2 
    for col_offset, header in enumerate(dept_headers):
        ws_overview.cell(row=1, column=dept_start_col + col_offset, value=header)

    loc_headers = ["Sim Name"] + [f"{l} % Comp (of Received)" for l in locations]
    loc_start_col = dept_start_col + len(dept_headers) + 2
    for col_offset, header in enumerate(loc_headers):
        ws_overview.cell(row=1, column=loc_start_col + col_offset, value=header)

    row_num = 2
    for sim_name, rows in sims.items():
        sim_date = rows[0][idx["Sim Date"]]
        complexity = rows[0][idx["Complexity"]]
        total = len(rows)
        received = sum(1 for r in rows if r[idx["Received"]])
        forwarded = sum(1 for r in rows if r[idx["Forwarded"]])
        read = sum(1 for r in rows if r[idx["Read"]])
        deleted = sum(1 for r in rows if r[idx["Deleted"]])
        reported = sum(1 for r in rows if r[idx["Reported"]])
        compromised = sum(1 for r in rows if r[idx["Compromised"]])

        pct_comp_delivered = round((compromised / total) * 100, 1) if total else 0
        pct_comp_read = round((compromised / read) * 100, 1) if read else 0

        core_values = [
            sim_name, sim_date, complexity,
            total, received, forwarded,
            read, deleted, reported,
            compromised, pct_comp_delivered, pct_comp_read
        ]
        for col_offset, val in enumerate(core_values):
            ws_overview.cell(row=row_num, column=core_start_col + col_offset, value=val)

        dept_values = [sim_name]
        for dept in departments:
            dept_rows = [r for r in rows if r[idx["Department"]] == dept]
            dept_received = [r for r in dept_rows if r[idx["Received"]]]
            if dept_received:
                comp_count = sum(1 for r in dept_received if r[idx["Compromised"]])
                dept_values.append(round((comp_count / len(dept_received)) * 100, 1))
            else:
                dept_values.append(0)
        for col_offset, val in enumerate(dept_values):
            ws_overview.cell(row=row_num, column=dept_start_col + col_offset, value=val)

        loc_values = [sim_name]
        for loc in locations:
            loc_rows = [r for r in rows if r[idx["Office Location"]] == loc]
            loc_received = [r for r in loc_rows if r[idx["Received"]]]
            if loc_received:
                comp_count = sum(1 for r in loc_received if r[idx["Compromised"]])
                loc_values.append(round((comp_count / len(loc_received)) * 100, 1))
            else:
                loc_values.append(0)
        for col_offset, val in enumerate(loc_values):
            ws_overview.cell(row=row_num, column=loc_start_col + col_offset, value=val)

        row_num += 1

    core_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    dept_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    loc_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for col in range(core_start_col, core_start_col + len(core_headers)):
        ws_overview.cell(row=1, column=col).font = Font(bold=True)
        ws_overview.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws_overview.cell(row=1, column=col).fill = core_fill

    for col in range(dept_start_col, dept_start_col + len(dept_headers)):
        ws_overview.cell(row=1, column=col).font = Font(bold=True)
        ws_overview.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws_overview.cell(row=1, column=col).fill = dept_fill

    for col in range(loc_start_col, loc_start_col + len(loc_headers)):
        ws_overview.cell(row=1, column=col).font = Font(bold=True)
        ws_overview.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws_overview.cell(row=1, column=col).fill = loc_fill

    for col in ws_overview.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_overview.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(file_path)
    print(f"Simulation Overview sheet updated in {file_path}")

def data_clean(file_path="excel/phishing_sim_data.xlsx"):
    dept_map = {
        # Accounting
        "accounts": "Accounting",
        "finance": "Accounting",
        "finance & accounting": "Accounting",

        # Administration
        "administration": "Administration",

        # Area Management
        "area management": "Area Management",

        # Assembly
        "assembly": "Assembly",
        "assembly & fab": "Assembly",

        # Construction
        "construction": "Construction",
        "shovel ready": "Construction",

        # Engineering
        "engineering": "Engineering",
        "controls": "Engineering",
        "design engineers - lebanon": "Engineering",
        "eam engineers": "Engineering",

        # Customer Support
        "customer support": "Customer Support",
        "office support": "Customer Support",
        "support": "Customer Support",
        "support services": "Customer Support",
        "customer service": "Customer Support",

        # Sales & Marketing
        "sales & marketing": "Sales & Marketing",
        "eam sales": "Sales & Marketing",
        "extruder sales": "Sales & Marketing",
        "sales and marketing": "Sales & Marketing",
        "specialty products": "Sales & Marketing",

        # IT
        "it": "IT",
        "emi it": "IT",
        "information technology": "IT",
        "microsoft communication application instance": "IT",
        "mis": "IT",
        "is": "IT",

        # Electrical Shop
        "electrical shop": "Electrical Shop",

        # Environmental
        "environmental": "Environmental",
        "environment": "Environmental",
        "environmental, health & safety": "Environmental",

        # Safety
        "safety": "Safety",

        # Management
        "management": "Management",
        "general affairs": "Management",
        "general management": "Management",
        "logistics": "Management",
        "management - executive consultant": "Management",
        "purchasing": "Management",
        "supervision": "Management",

        # Human Resources
        "human resources": "Human Resources",
        "hr": "Human Resources",
        "human resource/shared services": "Human Resources",


        # Machine Shop
        "machine shop": "Machine Shop",
        "fab shop": "Machine Shop",
        "machine division - assembly": "Machine Shop",
        "machine division - assembly nv": "Machine Shop",
        "machine division - controls": "Machine Shop",
        "machine division - controls nv": "Machine Shop",
        "machine division - electrical": "Machine Shop",
        "machine division - electrical nv": "Machine Shop",
        "machine division - lab": "Machine Shop",
        "machine division - machine shop": "Machine Shop",
        "machine division - machine shop nv": "Machine Shop",
        "machine division - mech engineer nv": "Machine Shop",
        "machine division - mechanical engineer": "Machine Shop",
        "machine division - operations": "Machine Shop",
        "machine division - operations (gen abs)": "Machine Shop",
        "machine division - operations nv": "Machine Shop",
        "machine division - rolls": "Machine Shop",
        "machine division - s & m": "Machine Shop",
        "machine division - s, g & a": "Machine Shop",

        # Maintenance
        "maintenance": "Maintenance",
        "installation": "Maintenance",
        "material handling field install": "Maintenance",

        # Production
        "production": "Production",
        "production management": "Production",
        "production support": "Production",
        "productions": "Production",
        "sli production": "Production",
        "teklon production": "Production",
        "prodcution": "Production",

        #Sales & Marketing
        "sales": "Sales & Marketing",

        # Quality
        "quality": "Quality",
        "qa": "Quality",
        "qc lab": "Quality",
        "quality assurance": "Quality",
        "qc": "Quality",

        # R&D
        "r&d": "R&D",
        "extruder lab": "R&D",
        "r & d personnel": "R&D",
        "technical development": "R&D",
        "technical development and global quality": "R&D",

        # Security
        "security": "Security",

        # Shipping & Receiving
        "shipping & receiving": "Shipping & Receiving",
        "shipping": "Shipping & Receiving",
        "shipping/receiving": "Shipping & Receiving",
        "warehouse": "Shipping & Receiving",

        # Supply Chain
        "supply chain": "Supply Chain",

        # Site Lead
        "site lead": "Site Lead",
        "eam": "Site Lead",
        "entek asia": "Site Lead",
        "entek processing": "Site Lead",
        "epi tsu plant": "Site Lead",
        "hse": "Site Lead",
        "mis department": "Site Lead",
        "nss": "Site Lead",
        "saratoga": "Site Lead",
        "tarui plant": "Site Lead",
        "teklon": "Site Lead",
        "tianjin plant": "Site Lead",
        "scm": "Site Lead",
    }

    location_map = {
        "bogor": "Bogor",
        "entek bogor": "Bogor",
        "lebanon": "Lebanon",
        "entek uk": "Newcastle",
        "henderson": "Henderson",
        "tarui": "Tarui",
        "entek tarui": "Tarui",
        "remote": "Remote"
    }

    wb = load_workbook(file_path)
    ws = wb["Phishing Simulation Data"]

    headers = [cell.value for cell in ws[1]]
    dept_idx = headers.index("Department")
    loc_idx = headers.index("Office Location")

    for row in ws.iter_rows(min_row=2):
        dept_val = row[dept_idx].value
        if dept_val:
            dept_key = str(dept_val).strip().lower()
            if dept_key in dept_map:
                row[dept_idx].value = dept_map[dept_key]

        loc_val = row[loc_idx].value
        if loc_val:
            loc_key = str(loc_val).strip().lower()
            if loc_key in location_map:
                row[loc_idx].value = location_map[loc_key]
            else:
                row[loc_idx].value = "Remote"

    wb.save(file_path)
    print(f"Data cleaned and saved to {file_path}")

if __name__ == "__main__":
    data_clean("excel/phishing_sim_data.xlsx")
    generate_simulation_overview("excel/phishing_sim_data.xlsx")
